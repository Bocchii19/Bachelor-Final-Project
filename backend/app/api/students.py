"""Students API — CRUD, Excel import, face enrollment."""

from __future__ import annotations

import os
import uuid
from typing import List, Optional

import numpy as np
import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import delete, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import get_current_user
from app.config import get_settings
from app.database import get_db
from app.models.face_embedding import FaceEmbedding
from app.models.student import Student
from app.models.user import User
from app.schemas.student import (
    EnrollmentResult,
    ImportResult,
    StudentCreate,
    StudentRead,
    StudentUpdate,
)

router = APIRouter()
settings = get_settings()

# Expected Excel column names (flexible mapping)
EXPECTED_COLUMNS = {
    "MSSV": "student_code",
    "Mã SV": "student_code",
    "Student Code": "student_code",
    "Họ và tên": "full_name",
    "Họ tên": "full_name",
    "Full Name": "full_name",
    "Email": "email",
}


# ── CRUD ──────────────────────────────────────────────

@router.get("/", response_model=List[StudentRead])
async def list_students(
    class_id: Optional[uuid.UUID] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """List students, optionally filter by class."""
    stmt = select(Student)
    if class_id:
        stmt = stmt.where(Student.class_id == class_id)
    stmt = stmt.order_by(Student.student_code).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("/", response_model=StudentRead, status_code=status.HTTP_201_CREATED)
async def create_student(
    body: StudentCreate,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    student = Student(**body.model_dump())
    db.add(student)
    await db.flush()
    await db.refresh(student)
    return student


@router.get("/{student_id}", response_model=StudentRead)
async def get_student(
    student_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.patch("/{student_id}", response_model=StudentRead)
async def update_student(
    student_id: uuid.UUID,
    body: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    for key, value in body.model_dump(exclude_unset=True).items():
        setattr(student, key, value)

    await db.flush()
    await db.refresh(student)
    return student


# ── Import from Excel ──────────────────────────────────

@router.post("/import", response_model=ImportResult)
async def import_students_from_excel(
    class_id: uuid.UUID = Query(..., description="Target class ID"),
    file: UploadFile = File(..., description="Excel file (.xlsx)"),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    Import students from an Excel file.
    Expected columns: MSSV | Họ và tên | Email
    Uses upsert (ON CONFLICT) to handle duplicates.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()

    # Parse Excel
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel file has no active sheet")

    # Map header row
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map: dict[int, str] = {}
    for idx, header in enumerate(headers):
        for expected, field in EXPECTED_COLUMNS.items():
            if header.lower() == expected.lower():
                col_map[idx] = field
                break

    if "student_code" not in col_map.values() or "full_name" not in col_map.values():
        raise HTTPException(
            status_code=400,
            detail=f"Excel must have columns: MSSV, Họ và tên. Found: {headers}",
        )

    # Process rows
    result = ImportResult()
    rows_to_upsert: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        result.total_rows += 1
        row_data: dict = {"class_id": class_id}

        for idx, field in col_map.items():
            value = row[idx] if idx < len(row) else None
            row_data[field] = str(value).strip() if value else None

        # Validate
        if not row_data.get("student_code"):
            result.errors.append(f"Row {result.total_rows + 1}: Missing MSSV")
            continue
        if not row_data.get("full_name"):
            result.errors.append(f"Row {result.total_rows + 1}: Missing name")
            continue

        rows_to_upsert.append(row_data)

    wb.close()

    # Bulk upsert
    if rows_to_upsert:
        stmt = pg_insert(Student).values(rows_to_upsert)
        stmt = stmt.on_conflict_do_update(
            index_elements=["student_code"],
            set_={
                "full_name": stmt.excluded.full_name,
                "email": stmt.excluded.email,
                "class_id": stmt.excluded.class_id,
            },
        )
        await db.execute(stmt)

        # Count inserted vs updated (approximate)
        existing = await db.execute(
            select(Student.student_code).where(
                Student.student_code.in_([r["student_code"] for r in rows_to_upsert])
            )
        )
        existing_codes = {r[0] for r in existing.all()}
        result.updated = len(existing_codes)
        result.inserted = len(rows_to_upsert) - result.updated

    return result


# ── Face Enrollment ──────────────────────────────────

@router.post("/{student_id}/enroll-face", response_model=EnrollmentResult)
async def enroll_face(
    student_id: uuid.UUID,
    images: List[UploadFile] = File(..., description="3-5 face images from different angles"),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    Enroll a student's face.
    1. Receive 3–5 images from different angles.
    2. Detect face (reject if 0 or >1 faces per image).
    3. Liveness check.
    4. Compute ArcFace embedding (512-dim).
    5. Save ALL embeddings (not averaged) for max-similarity matching.
    6. Save cropped face images to storage.
    """
    # Verify student exists
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if len(images) < 1:
        raise HTTPException(status_code=400, detail="At least 1 image required")
    if len(images) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 images allowed")

    # Lazy import CV modules
    from app.cv.pipeline import CVPipeline
    pipeline = CVPipeline.get_instance()

    enrollment = EnrollmentResult(
        student_id=student_id,
        embeddings_created=0,
        images_saved=[],
    )

    save_dir = os.path.join(settings.MEDIA_ROOT, "faces", str(student_id))
    os.makedirs(save_dir, exist_ok=True)

    for idx, img_file in enumerate(images):
        try:
            # Read image
            img_bytes = await img_file.read()
            img_array = np.frombuffer(img_bytes, dtype=np.uint8)

            import cv2
            frame = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
            if frame is None:
                enrollment.errors.append(f"Image {idx}: Could not decode")
                continue

            # Detect faces
            faces = pipeline.detect_faces(frame)
            if len(faces) == 0:
                enrollment.errors.append(f"Image {idx}: No face detected")
                continue
            if len(faces) > 1:
                enrollment.errors.append(f"Image {idx}: Multiple faces detected ({len(faces)})")
                continue

            face = faces[0]

            # Liveness check
            is_live = pipeline.check_liveness(frame, face)
            if not is_live:
                enrollment.errors.append(f"Image {idx}: Failed liveness check (possible spoof)")
                continue

            # Compute embedding
            embedding = pipeline.compute_embedding(frame, face)
            if embedding is None:
                enrollment.errors.append(f"Image {idx}: Could not compute embedding")
                continue

            # Save cropped face image
            bbox = face.bbox.astype(int)
            x1, y1, x2, y2 = max(0, bbox[0]), max(0, bbox[1]), bbox[2], bbox[3]
            face_crop = frame[y1:y2, x1:x2]
            filename = f"face_{idx}.jpg"
            filepath = os.path.join(save_dir, filename)
            cv2.imwrite(filepath, face_crop)

            # Save embedding to DB
            db_embedding = FaceEmbedding(
                student_id=student_id,
                embedding=embedding.tolist(),
                image_path=filepath,
            )
            db.add(db_embedding)
            enrollment.embeddings_created += 1
            enrollment.images_saved.append(filepath)

        except Exception as e:
            enrollment.errors.append(f"Image {idx}: {str(e)}")
            continue

    await db.flush()
    return enrollment


# ── Delete Embedding ──────────────────────────────────

@router.delete("/{student_id}/embedding", status_code=status.HTTP_204_NO_CONTENT)
async def delete_embeddings(
    student_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Delete all face embeddings for a student."""
    await db.execute(
        delete(FaceEmbedding).where(FaceEmbedding.student_id == student_id)
    )
