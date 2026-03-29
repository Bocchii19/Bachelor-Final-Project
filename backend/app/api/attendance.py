"""Attendance API — Pivot sheet, export Excel."""

from __future__ import annotations

import io
import uuid
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import get_current_user
from app.database import get_db
from app.models.attendance import AttendanceRecord
from app.models.class_ import Class
from app.models.session import Session
from app.models.student import Student
from app.models.user import User
from app.schemas.attendance import (
    AttendanceCell,
    AttendanceSheetData,
    SessionColumn,
    StudentRow,
)

router = APIRouter()


@router.get("/sheet", response_model=AttendanceSheetData)
async def get_attendance_sheet(
    class_id: uuid.UUID = Query(...),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """
    Build pivot attendance sheet data.
    Rows = students, Columns = session dates, Cells = attendance status.
    """
    # Fetch class info
    cls_result = await db.execute(select(Class).where(Class.id == class_id))
    cls = cls_result.scalar_one_or_none()
    if not cls:
        raise HTTPException(status_code=404, detail="Class not found")

    # Fetch sessions (columns)
    sess_stmt = select(Session).where(Session.class_id == class_id)
    if date_from:
        sess_stmt = sess_stmt.where(Session.session_date >= date_from)
    if date_to:
        sess_stmt = sess_stmt.where(Session.session_date <= date_to)
    sess_stmt = sess_stmt.order_by(Session.session_date)

    sessions_result = await db.execute(sess_stmt)
    sessions = sessions_result.scalars().all()

    # Fetch students (rows)
    students_result = await db.execute(
        select(Student)
        .where(Student.class_id == class_id)
        .order_by(Student.student_code)
    )
    students = students_result.scalars().all()

    if not sessions:
        return AttendanceSheetData(
            class_id=class_id,
            class_name=cls.name,
            subject=cls.subject,
            columns=[],
            rows=[
                StudentRow(
                    student_id=s.id,
                    student_code=s.student_code,
                    full_name=s.full_name,
                    attendance={},
                )
                for s in students
            ],
        )

    session_ids = [s.id for s in sessions]

    # Fetch all attendance records for these sessions
    records_result = await db.execute(
        select(AttendanceRecord).where(
            AttendanceRecord.session_id.in_(session_ids)
        )
    )
    records = records_result.scalars().all()

    # Build lookup: (student_id, session_id) → record
    record_map: dict[tuple[uuid.UUID, uuid.UUID], AttendanceRecord] = {}
    for r in records:
        record_map[(r.student_id, r.session_id)] = r

    # Build columns
    columns: list[SessionColumn] = []
    for sess in sessions:
        present = sum(
            1
            for r in records
            if r.session_id == sess.id and r.status == "present"
        )
        unknown = sum(
            1
            for r in records
            if r.session_id == sess.id and r.status == "unknown"
        )
        columns.append(
            SessionColumn(
                session_id=sess.id,
                session_date=sess.session_date,
                status=sess.status,
                present_count=present,
                unknown_count=unknown,
                total_students=len(students),
            )
        )

    # Build rows
    rows: list[StudentRow] = []
    for student in students:
        attendance: dict[str, AttendanceCell] = {}
        present_count = 0

        for sess in sessions:
            key = (student.id, sess.id)
            rec = record_map.get(key)
            if rec:
                attendance[sess.session_date.isoformat()] = AttendanceCell(
                    status=rec.status,
                    confidence=rec.confidence,
                )
                if rec.status == "present":
                    present_count += 1
            else:
                attendance[sess.session_date.isoformat()] = AttendanceCell(
                    status="absent"
                )

        total = len(sessions)
        rows.append(
            StudentRow(
                student_id=student.id,
                student_code=student.student_code,
                full_name=student.full_name,
                attendance=attendance,
                present_count=present_count,
                total_sessions=total,
                attendance_rate=round(present_count * 100.0 / total, 1) if total else 0.0,
            )
        )

    return AttendanceSheetData(
        class_id=class_id,
        class_name=cls.name,
        subject=cls.subject,
        columns=columns,
        rows=rows,
    )


@router.get("/export")
async def export_attendance_excel(
    class_id: uuid.UUID = Query(...),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Export attendance sheet as .xlsx file."""
    # Reuse sheet data
    sheet = await get_attendance_sheet(class_id, date_from, date_to, db, _user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header row
    headers = ["MSSV", "Họ và tên"]
    for col in sheet.columns:
        headers.append(col.session_date.strftime("%d/%m/%Y"))
    headers.extend(["Có mặt", "Tổng buổi", "Tỉ lệ (%)"])
    ws.append(headers)

    # Style header
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    status_fills = {
        "present": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "absent": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "unknown": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }
    status_symbols = {"present": "✓", "absent": "—", "unknown": "?"}

    for row in sheet.rows:
        row_data = [row.student_code, row.full_name]
        for col in sheet.columns:
            cell_data = row.attendance.get(col.session_date.isoformat())
            if cell_data:
                row_data.append(status_symbols.get(cell_data.status, cell_data.status))
            else:
                row_data.append("—")
        row_data.extend([row.present_count, row.total_sessions, row.attendance_rate])
        ws.append(row_data)

    # Apply status colors to data cells
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(3, 3 + len(sheet.columns)):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            cell.alignment = Alignment(horizontal="center")
            if value == "✓":
                cell.fill = status_fills["present"]
            elif value == "—":
                cell.fill = status_fills["absent"]
            elif value == "?":
                cell.fill = status_fills["unknown"]

    # Auto-width columns
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 10)

    # Summary row
    ws.append([])
    summary_row = ["", "Tổng có mặt"]
    for col in sheet.columns:
        summary_row.append(col.present_count)
    ws.append(summary_row)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{sheet.class_name}_{date_from or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
